from openpyxl import *
import datetime
import os

import excelstyle


def mergeExcelData(setString):
    fullString = ''
    data = ''
    for key, value in setString.items():
        if '_PS' in key:
            data = f'[ {key.split("_")[0]} 프로세스 상태 ]\n'
        elif '_PORT' in key:
            data = f'[ {key.split("_")[0]} 포트 상태 ]\n'
        elif '_SYS' in key:
            data = f'[ {key.split("_")[0]} 서비스 데몬 상태 ]\n'
        elif 'FILEPERM:' in key:
            data = f'파일명: {key.split("FILEPERM:")[1]}\n'
        elif 'FILEDATA:' in key:
            data = f'파일명: {key.split("FILEDATA:")[1]}\n'
        elif 'CMD:' in key:
            data = f'[ {key.split("CMD:")[1]} ]\n\n'
        data += f'{value}\n'
        fullString += data

    return fullString


def makeExcelReport(analysisRes, sysList):
    dt = datetime.datetime.now()

    wb = load_workbook('default_report.xlsx')
    wsResSum = wb['진단 결과 요약']
    impDict = {}
    totalCnt = 0    # 진단한 전체 항목 수
    resultCnt = 0   # 양호 또는 리뷰에 대한 합산 점수

    # 자산 정보
    wsResSum['D6'] = sysList['osType']
    wsResSum['D7'] = f'{sysList["osName"]} {sysList["osVersion"]}'
    wsResSum['D8'] = sysList['hostname']
    wsResSum['D9'] = dt.strftime("%Y-%m-%d %H:%M:%S")

    # 항목 별로 결과에 따라 점수 자체를 넣는 방식
    # 추후 상태 및 중요도 별 분포표 작성시에 사용
    # {'양호': [3, 3, 1], '취약': [3, 2], '리뷰': [3, 3]}
    impDict.update({'양호': [int(data[3]['ImportantScore']) for data in analysisRes if data[1] in '양호' ]})
    impDict.update({'취약': [int(data[3]['ImportantScore']) for data in analysisRes if data[1] in '취약']})
    impDict.update({'리뷰': [int(data[3]['ImportantScore']) for data in analysisRes if data[1] in '리뷰']})

    # 상태 별 결과 분포표
    for cnt, val in zip(range(0, 4), [len(analysisRes), len(impDict['양호']), len(impDict['취약']), len(impDict['리뷰'])]):
        wsResSum.cell(row=19, column=10 + cnt).font = excelstyle.normalfont
        wsResSum.cell(row=19, column=10 + cnt).value = val

    # 중요도 별 결과 분포표
    for colcnt, chkval in zip(range(0, 3), ['양호', '취약', '리뷰']):
        for rowcnt in range(0, 3):
            valcnt = impDict[chkval].count(rowcnt + 1)
            wsResSum.cell(row=29 + rowcnt, column=11 + colcnt).font = excelstyle.normalfont
            wsResSum.cell(row=29 + rowcnt, column=11 + colcnt).value = valcnt
            totalCnt += wsResSum.cell(row=29 + rowcnt, column=10).value * valcnt
            if colcnt == 0:
                resultCnt += wsResSum.cell(row=29 + rowcnt, column=10).value * valcnt
            elif colcnt == 2:
                resultCnt += wsResSum.cell(row=29 + rowcnt, column=10).value * valcnt * 0.5

    wsResSum['D2'] = resultCnt / totalCnt * 100
    wsResSum['D2'].font = excelstyle.headredfont

    # 진단 결과 내역
    # 항목 구분 기준으로 정렬 후 내역에 삽입
    analysisRes.sort(key=lambda x: x[3]['Category'])
    # 총 항목 개수만큼만 반복을 돌면서 아래로 항목 하나씩 삽입
    for cnt in range(0, len(analysisRes)):
        # 진단 결과 중 필요한 부문만 뽑아내어 사용
        inputdata = [analysisRes[cnt][3]['Category'], analysisRes[cnt][0], analysisRes[cnt][3]['Name'],
                     int(analysisRes[cnt][3]['ImportantScore']), analysisRes[cnt][1]]
        # 병합되어진 셀 들의 위치를 맞추기 위해 column 부분을 띄어줌
        for num, idx in zip([0, 2, 3, 10, 11], range(0, len(inputdata))):
            # 각 부분에 맞게 왼쪽 정렬 또는 오른족 정렬 및 진단 결과에 따라 글자 색 변경
            wsResSum.cell(row=36+cnt, column=2+num).font = excelstyle.normalfont
            if num == 10 or num == 11:
                wsResSum.cell(row=36 + cnt, column=2 + num).alignment = excelstyle.centeralign

            if num == 11:
                if inputdata[idx] == '양호':
                    wsResSum.cell(row=36 + cnt, column=2 + num).font = excelstyle.greenfont
                elif inputdata[idx] == '리뷰':
                    wsResSum.cell(row=36 + cnt, column=2 + num).font = excelstyle.bluefont
                elif inputdata[idx] == '취약':
                    wsResSum.cell(row=36 + cnt, column=2 + num).font = excelstyle.redfont

            # 모든 칸에 테두리 적용 및 알맞은 결과 값 입력
            wsResSum.cell(row=36 + cnt, column=2 + num).border = excelstyle.thinborder
            wsResSum.cell(row=36 + cnt, column=2 + num).value = inputdata[idx]
            # 표에서 항목코드 다음으로 들어가는 항목명이 들어가는 부분 병합
            if num == 3:
                wsResSum.merge_cells(start_row=36 + cnt, start_column=2 + num, end_row=36 + cnt, end_column=2 + num + 6)
            # 표에서 구분 부분의 구분명 들어가는 부분 병합
            elif num == 0:
                wsResSum.merge_cells(start_row=36 + cnt, start_column=2 + num, end_row=36 + cnt, end_column=2 + num + 1)

    # 진단 결과 상세 내역(세로)
    wsResDetVer = wb['진단 결과 상세']
    # 세번째 줄부터 입력 시작
    rownum = 3
    for cnt in range(0, len(analysisRes)):
        inputdata = [
            ['', ''], ['구분', analysisRes[cnt][3]['Category']], ['코드', analysisRes[cnt][0]],
            ['항목', analysisRes[cnt][3]['Name']], ['중요도', int(analysisRes[cnt][3]['ImportantScore'])],
            ['진단 결과', analysisRes[cnt][1]], ['판단 기준', analysisRes[cnt][3]['Criterion']],
            ['상세 현황', analysisRes[cnt][2]], ['조치 방법', analysisRes[cnt][3]['ActionPlan']]
        ]
        # inputdata 내용 index 값
        for idx in range(0, 9):
            for colcnt in range(0, 2):
                # 표 시작 전 줄은 빈 줄이기 때문에 inputdata의 index 값이 0이 아닐 경우에만 테두리 지정
                if idx != 0:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).border = excelstyle.thinborder
                if colcnt == 0 and idx != 0:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).alignment = excelstyle.centerwrapalign
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).fill = excelstyle.cellbgfill
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).font = excelstyle.whiteboldfont
                else:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).alignment = excelstyle.leftwrapalign
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).font = excelstyle.normalfont

                if idx == 7 and colcnt == 1:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).value = mergeExcelData(inputdata[idx][colcnt])
                else:
                    wsResDetVer.cell(row=rownum, column=2 + colcnt).value = inputdata[idx][colcnt]

            rownum += 1

    # 진단 결과 상세내역(가로)
    wsResDetHor = wb['진단 결과 상세(가로)']
    for cnt in range(0, len(analysisRes)):
        inputdata = [
            analysisRes[cnt][3]['Category'], analysisRes[cnt][0], analysisRes[cnt][3]['Name'],
            int(analysisRes[cnt][3]['ImportantScore']), analysisRes[cnt][1], analysisRes[cnt][3]['Criterion'],
            analysisRes[cnt][2], analysisRes[cnt][3]['ActionPlan']
        ]
        for idx in range(0, 8):
            wsResDetHor.cell(row=5 + cnt, column=2 + idx).border = excelstyle.thinborder
            wsResDetHor.cell(row=5 + cnt, column=2 + idx).font = excelstyle.normalfont
            if idx == 6:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).value = mergeExcelData(inputdata[idx])
            else:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).value = inputdata[idx]
            if idx < 2:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).alignment = excelstyle.leftalign
            elif idx == 4 or idx == 3:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).alignment = excelstyle.centeralign
                if inputdata[idx] == '양호':
                    wsResDetHor.cell(row=5 + cnt, column=2 + idx).font = excelstyle.greenfont
                elif inputdata[idx] == '리뷰':
                    wsResDetHor.cell(row=5 + cnt, column=2 + idx).font = excelstyle.bluefont
                if inputdata[idx] == '취약':
                    wsResDetHor.cell(row=5 + cnt, column=2 + idx).font = excelstyle.redfont
            else:
                wsResDetHor.cell(row=5 + cnt, column=2 + idx).alignment = excelstyle.fillalign

    # 참고 시트
    wsResRefInfo = wb['참고']
    # 해당 시트에는 네트워크 정보, 프로세스 전체 내역, 포트 전체 내역, 서비스 내역을 참고사항으로 입력
    for info, val in zip(['ipList', 'processInfo', 'portInfo', 'serviceInfo'], ['B5', 'B31', 'B57', 'B83']):
        if info in sysList.keys():
            wsResRefInfo[val] = sysList[info]
            wsResRefInfo[val].font = excelstyle.normalfont
            wsResRefInfo[val].alignment = excelstyle.leftwrapalign

    fileName = f'result_report_{sysList["osName"]}_{sysList["hostname"]}_{dt.strftime("%Y%m%d%H%M%S")}.xlsx'

    wb.save(filename=os.path.join(os.getcwd(), 'ExcelDir', fileName))

    return fileName
